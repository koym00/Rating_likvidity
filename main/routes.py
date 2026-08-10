import csv
import datetime
import io
import math
import unicodedata
import zipfile

import openpyxl
import requests
from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for

bp_main = Blueprint("main", __name__, template_folder="templates", static_folder="static")

# --- živé zdroje dat (ČÚZK / ČSÚ) -------------------------------------------------
CUZK_VKLADY_URL_VZOR = (
    "https://cuzk.gov.cz/Periodika-a-publikace/Statisticke-udaje/"
    "Statistiky-podle-roku/{rok}/statistiky-{rok}.aspx"
)
CUZK_VKLADY_ENTRY_VZOR = "statistiky-{rok}-1234Q-V.xlsx"
POCET_LET_ZPET_KE_ZKUSENI = 2  # kolik let dozadu zkusit, když aktuální rok ještě nemá kompletní data

CUZK_CISELNIK_ZIP_URL = "https://services.cuzk.cz/sestavy/cis/SC_SEZNAMKUKRA_DOTAZ.zip"
CUZK_CISELNIK_ENTRY = "SC_SEZNAMKUKRA_DOTAZ.csv"

# Oficiální DataStat API ČSÚ (stejný zdroj dat jako widget na csu.gov.cz/regionalni-statistiky,
# tabulka "Počet obyvatel k 31.12."). Vrací strukturovaná JSON-stat data se všemi kraji a lety
# najednou, takže se vždy automaticky použije nejnovější dostupný rok - žádné PDF ani ruční
# aktualizace odkazu není potřeba.
CSU_OBYVATELE_API_URL = (
    "https://csu.gov.cz/datastat/api/dotaz/data/vybery/uzivatelske/"
    "59583b40-c6b0-42db-b5b1-83d84eb10143?jazyk=cs"
)

CANONICAL_KRAJE = [
    "Středočeský", "Hlavní město Praha", "Jihomoravský", "Moravskoslezský",
    "Ústecký", "Jihočeský", "Olomoucký", "Plzeňský", "Zlínský",
    "Královéhradecký", "Pardubický", "Vysočina", "Liberecký", "Karlovarský",
]


def normalize_kraj(name):
    if not name:
        return ""
    name = str(name).replace("Kraj ", "").replace(" kraj", "").strip()
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return name.lower()


def parse_statistika(file_stream):
    """Načte hárok se statistikou vkladů a vrátí seznam krajů s hodnotou
    'vlastnické právo, celkem'."""
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    col_vlastnicke = None
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if not val:
                continue
            if "vlastnick" in str(val).lower():
                col_vlastnicke = c
                break
        if col_vlastnicke:
            break

    if col_vlastnicke is None:
        raise ValueError("V souboru se statistikou nebyl nalezen sloupec 'vlastnické právo'.")

    kraje = []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if name is None:
            if kraje:
                break
            continue
        if b not in (None, ""):
            break  # narazili jsme na rozpis podle pracovišť -> konec bloku krajů

        d = ws.cell(row=r, column=col_vlastnicke).value
        if d is None:
            continue

        is_total = "republik" in str(name).lower()
        kraje.append({"name": str(name).strip(), "D": float(d), "is_total": is_total})

    if not kraje:
        raise ValueError("V souboru se statistikou se nepodařilo najít žádné řádky krajů.")

    return kraje


def detect_encoding(raw_bytes):
    # UTF-8 musí být zkoušeno první: cp1250 je jednobajtové kódování, které téměř
    # jakýkoli bajtový vstup "úspěšně" dekóduje (jen s nesmyslným výsledkem), takže
    # kdyby šlo první, reálně UTF-8 soubory by se nikdy nesprávně nerozpoznaly.
    for enc in ("utf-8-sig", "utf-8", "cp1250"):
        try:
            raw_bytes.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "cp1250"


def parse_obyvatele_csv(raw_bytes):
    enc = detect_encoding(raw_bytes)
    text = raw_bytes.decode(enc)
    delim = ";" if text.count(";") >= text.count(",") else ","

    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        raise ValueError("Soubor s počtem obyvatel je prázdný.")

    header = [str(h).strip().lower() for h in rows[0]]
    name_idx = next((i for i, h in enumerate(header) if "kraj" in h), None)
    pop_idx = next((i for i, h in enumerate(header) if "obyvatel" in h), None)
    if name_idx is None or pop_idx is None:
        raise ValueError(
            "Soubor s počtem obyvatel musí mít v hlavičce sloupec s názvem kraje "
            "(obsahující slovo 'kraj') a sloupec s počtem obyvatel (obsahující 'obyvatel')."
        )

    population = {}
    for row in rows[1:]:
        if len(row) <= max(name_idx, pop_idx):
            continue
        name = str(row[name_idx]).strip()
        raw_pop = str(row[pop_idx]).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        if not name or not raw_pop:
            continue
        try:
            population[normalize_kraj(name)] = float(raw_pop)
        except ValueError:
            continue

    if not population:
        raise ValueError("Ze souboru s počtem obyvatel se nepodařilo načíst žádné řádky.")
    return population


def parse_obyvatele_xlsx(file_stream):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    header = [str(ws.cell(row=1, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
    name_col = next((i + 1 for i, h in enumerate(header) if "kraj" in h), None)
    pop_col = next((i + 1 for i, h in enumerate(header) if "obyvatel" in h), None)
    if name_col is None or pop_col is None:
        raise ValueError(
            "Soubor s počtem obyvatel musí mít v hlavičce sloupec s názvem kraje "
            "(obsahující slovo 'kraj') a sloupec s počtem obyvatel (obsahující 'obyvatel')."
        )

    population = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        pop = ws.cell(row=r, column=pop_col).value
        if name is None or pop is None:
            continue
        try:
            population[normalize_kraj(name)] = float(pop)
        except (TypeError, ValueError):
            continue

    if not population:
        raise ValueError("Ze souboru s počtem obyvatel se nepodařilo načíst žádné řádky.")
    return population


def parse_obyvatele(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm")):
        return parse_obyvatele_xlsx(file_storage.stream)
    return parse_obyvatele_csv(file_storage.read())


def compute_ratings(kraje, population):
    """Rating likvidity = (celostátní míra = SUM(D)/SUM(P)*1000) / míra kraje,
    zaokrouhleno vždy dolů na 2 desetinná místa."""
    regiony = []
    total_row = None
    nenaparovane_obyvatelia = set()
    for k in kraje:
        p = population.get(normalize_kraj(k["name"]))
        k["P"] = p
        if k["is_total"]:
            total_row = k
        elif p is None:
            nenaparovane_obyvatelia.add(k["name"])
        else:
            regiony.append(k)

    if nenaparovane_obyvatelia:
        raise ValueError(
            "Pro tyto kraje ze souboru se statistikou se nepodařilo najít počet obyvatel: "
            + ", ".join(sorted(nenaparovane_obyvatelia))
        )

    sum_d = sum(k["D"] for k in regiony)
    sum_p = sum(k["P"] for k in regiony)
    if sum_p == 0:
        raise ValueError("Součet počtu obyvatel vyšel 0.")
    narodni_mira = sum_d / sum_p * 1000

    # celostátní řádek nemusí mít vlastní počet obyvatel v souboru - dopočítáme ho jako součet krajů
    if total_row is not None and total_row["P"] is None:
        total_row["P"] = sum_p

    ratings = {}
    for k in regiony + ([total_row] if total_row is not None else []):
        if not k["P"]:
            continue
        mira_kraje = math.floor((k["D"] / k["P"]) * 1000 * 100) / 100
        if mira_kraje == 0:
            continue
        rating = math.floor((narodni_mira / mira_kraje) * 100) / 100
        ratings[normalize_kraj(k["name"])] = rating
    return ratings


def normalize_ku_kod(val):
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def process_csv_ciselnik(raw_bytes, ratings):
    enc = detect_encoding(raw_bytes)
    text = raw_bytes.decode(enc)
    delim = ";" if text.count(";") >= text.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        raise ValueError("Číselník je prázdný.")

    header = rows[0]
    if "KRAJ_NAZEV" not in header:
        raise ValueError("Číselník neobsahuje sloupec 'KRAJ_NAZEV', podle kterého se páruje kraj.")
    nazev_idx = header.index("KRAJ_NAZEV")
    ku_kod_idx = header.index("KU_KOD") if "KU_KOD" in header else None

    header = header + ["Rating likvidity"]
    out_rows = [header]
    nenaparovane = set()
    ku_kod_to_rating = {}
    for row in rows[1:]:
        if not row:
            continue
        key = normalize_kraj(row[nazev_idx]) if nazev_idx < len(row) else ""
        rating = ratings.get(key)
        if rating is None:
            nenaparovane.add(row[nazev_idx] if nazev_idx < len(row) else "?")
            row = row + [""]
        else:
            row = row + [str(rating).replace(".", ",")]
            if ku_kod_idx is not None and ku_kod_idx < len(row):
                ku_kod = normalize_ku_kod(row[ku_kod_idx])
                if ku_kod:
                    ku_kod_to_rating[ku_kod] = rating
        out_rows.append(row)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delim, lineterminator="\r\n")
    writer.writerows(out_rows)
    data = buf.getvalue().encode(enc)

    out_stream = io.BytesIO(data)
    out_stream.seek(0)
    return out_stream, "ciselnik_s_ratingem.csv", nenaparovane, ku_kod_to_rating


def process_xlsx_ciselnik(file_stream, ratings):
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active

    header_row = 1
    header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    if "KRAJ_NAZEV" not in header_vals:
        raise ValueError("Číselník neobsahuje sloupec 'KRAJ_NAZEV', podle kterého se páruje kraj.")
    nazev_col = header_vals.index("KRAJ_NAZEV") + 1
    ku_kod_col = header_vals.index("KU_KOD") + 1 if "KU_KOD" in header_vals else None

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col, value="Rating likvidity")

    nenaparovane = set()
    ku_kod_to_rating = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=nazev_col).value
        if name is None:
            continue
        key = normalize_kraj(name)
        rating = ratings.get(key)
        if rating is None:
            nenaparovane.add(name)
            continue
        cell = ws.cell(row=r, column=new_col, value=rating)
        cell.number_format = "0.00"
        if ku_kod_col is not None:
            ku_kod = normalize_ku_kod(ws.cell(row=r, column=ku_kod_col).value)
            if ku_kod:
                ku_kod_to_rating[ku_kod] = rating

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    return out_stream, "ciselnik_s_ratingem.xlsx", nenaparovane, ku_kod_to_rating


def process_aktualizace_xlsx(file_stream, ku_kod_to_rating):
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active

    header_row = 1
    header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    if "Katastrální území" not in header_vals:
        raise ValueError("Soubor k aktualizaci neobsahuje sloupec 'Katastrální území'.")
    if "Rating likvidity" not in header_vals:
        raise ValueError("Soubor k aktualizaci neobsahuje sloupec 'Rating likvidity'.")
    ku_col = header_vals.index("Katastrální území") + 1
    rating_col = header_vals.index("Rating likvidity") + 1

    aktualizovano = 0
    nenaparovane = set()
    for r in range(header_row + 1, ws.max_row + 1):
        ku_kod = normalize_ku_kod(ws.cell(row=r, column=ku_col).value)
        if not ku_kod:
            continue
        rating = ku_kod_to_rating.get(ku_kod)
        if rating is None:
            nenaparovane.add(ku_kod)
            continue
        cell = ws.cell(row=r, column=rating_col, value=rating)
        cell.number_format = "0.00"
        aktualizovano += 1

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    return out_stream, "aktualizovany_soubor.xlsx", nenaparovane, aktualizovano


def process_aktualizace_csv(raw_bytes, ku_kod_to_rating):
    enc = detect_encoding(raw_bytes)
    text = raw_bytes.decode(enc)
    delim = ";" if text.count(";") >= text.count(",") else ","

    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        raise ValueError("Soubor k aktualizaci je prázdný.")

    header = rows[0]
    if "Katastrální území" not in header:
        raise ValueError("Soubor k aktualizaci neobsahuje sloupec 'Katastrální území'.")
    if "Rating likvidity" not in header:
        raise ValueError("Soubor k aktualizaci neobsahuje sloupec 'Rating likvidity'.")
    ku_idx = header.index("Katastrální území")
    rating_idx = header.index("Rating likvidity")

    aktualizovano = 0
    nenaparovane = set()
    out_rows = [header]
    for row in rows[1:]:
        if not row:
            continue
        ku_kod = normalize_ku_kod(row[ku_idx]) if ku_idx < len(row) else None
        if ku_kod:
            rating = ku_kod_to_rating.get(ku_kod)
            if rating is None:
                nenaparovane.add(ku_kod)
            else:
                while len(row) <= rating_idx:
                    row.append("")
                row[rating_idx] = str(rating).replace(".", ",")
                aktualizovano += 1
        out_rows.append(row)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delim, lineterminator="\r\n")
    writer.writerows(out_rows)
    data = buf.getvalue().encode(enc)

    out_stream = io.BytesIO(data)
    out_stream.seek(0)
    return out_stream, "aktualizovany_soubor.csv", nenaparovane, aktualizovano


def process_aktualizace(file_storage, ku_kod_to_rating):
    filename = (file_storage.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm")):
        return process_aktualizace_xlsx(file_storage.stream, ku_kod_to_rating)
    return process_aktualizace_csv(file_storage.read(), ku_kod_to_rating)


@bp_main.route("/", methods=["GET"])
def index():
    return render_template("index.html")


def fetch_zip_entry(zip_url, entry_name, popis):
    try:
        resp = requests.get(zip_url, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise ValueError(f"Nepodařilo se stáhnout {popis} z internetu ({exc}).") from exc

    try:
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            with zf.open(entry_name) as f:
                return f.read()
    except (zipfile.BadZipFile, KeyError) as exc:
        raise ValueError(
            f"Stažený soubor ({popis}) nemá očekávaný formát nebo neobsahuje "
            f"položku '{entry_name}' ({exc})."
        ) from exc


def fetch_vklady_bytes():
    """Zkusí aktuální rok, a pokud pro něj ještě není kompletní roční (1.-4. Q)
    statistika vkladů, spadne na předchozí roky - takže se nikde nemusí ručně
    měnit rok v kódu."""
    aktualni_rok = datetime.date.today().year
    posledni_chyba = None

    for rok in range(aktualni_rok, aktualni_rok - POCET_LET_ZPET_KE_ZKUSENI - 1, -1):
        url = CUZK_VKLADY_URL_VZOR.format(rok=rok)
        entry = CUZK_VKLADY_ENTRY_VZOR.format(rok=rok)
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
                with zf.open(entry) as f:
                    return f.read()
        except (requests.RequestException, zipfile.BadZipFile, KeyError) as exc:
            posledni_chyba = f"rok {rok}: {exc}"
            continue

    raise ValueError(
        "Nepodařilo se najít kompletní roční statistiku vkladů (zkoušeny roky "
        f"{aktualni_rok} až {aktualni_rok - POCET_LET_ZPET_KE_ZKUSENI}). Poslední chyba: {posledni_chyba}"
    )


def fetch_obyvatele_csv_bytes():
    try:
        resp = requests.get(CSU_OBYVATELE_API_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        data = resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise ValueError(f"Nepodařilo se stáhnout počet obyvatel z API ČSÚ ({exc}).") from exc

    try:
        uz02 = data["dimension"]["Uz02"]["category"]
        casr = data["dimension"]["CasR"]["category"]
        values = data["value"]
        n_years = len(casr["index"])
        nejnovejsi_rok = max(casr["index"], key=lambda rok: int(rok))
        rok_idx = casr["index"][nejnovejsi_rok]

        population = {}
        for kod, uz_idx in uz02["index"].items():
            if kod == "CZ":
                continue  # celostátní součet - nepočítá se jako samostatný kraj
            nazev = uz02["label"][kod]
            population[normalize_kraj(nazev)] = values[uz_idx * n_years + rok_idx]
    except (KeyError, IndexError, TypeError) as exc:
        raise ValueError(f"API ČSÚ vrátilo neočekávanou strukturu dat ({exc}).") from exc

    chybejici = [c for c in CANONICAL_KRAJE if normalize_kraj(c) not in population]
    if chybejici:
        raise ValueError(
            "API ČSÚ nevrátilo počet obyvatel pro všechny kraje (chybí: " + ", ".join(chybejici) + ")."
        )

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["Kraj", "Počet obyvatel", "Rok"])
    for cname in CANONICAL_KRAJE:
        writer.writerow([cname, population[normalize_kraj(cname)], nejnovejsi_rok])
    return buf.getvalue().encode("utf-8-sig")


@bp_main.route("/stahnout/<klic>")
def stahnout(klic):
    try:
        if klic == "statistika":
            data = fetch_vklady_bytes()
            filename = "statistika_vkladu.xlsx"
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif klic == "obyvatele":
            data = fetch_obyvatele_csv_bytes()
            filename = "pocet_obyvatel_kraje.csv"
            mimetype = "text/csv"
        elif klic == "ciselnik":
            data = fetch_zip_entry(CUZK_CISELNIK_ZIP_URL, CUZK_CISELNIK_ENTRY, "číselník katastrálních území (ČÚZK)")
            filename = CUZK_CISELNIK_ENTRY
            mimetype = "text/csv"
        else:
            abort(404)
    except ValueError as exc:
        return str(exc), 502

    return send_file(io.BytesIO(data), as_attachment=True, download_name=filename, mimetype=mimetype)


@bp_main.route("/vypocitat", methods=["POST"])
def vypocitat():
    f_statistika = request.files.get("statistika")
    f_obyvatele = request.files.get("obyvatele")
    f_ciselnik = request.files.get("ciselnik")
    f_aktualizace = request.files.get("aktualizace")
    ma_aktualizaci = bool(f_aktualizace and f_aktualizace.filename)

    if not f_statistika or f_statistika.filename == "":
        flash("Nahrajte prosím soubor se statistikou vkladů (XLSX).")
        return redirect(url_for("main.index"))
    if not f_obyvatele or f_obyvatele.filename == "":
        flash("Nahrajte prosím soubor s počtem obyvatel podle krajů.")
        return redirect(url_for("main.index"))
    if not f_ciselnik or f_ciselnik.filename == "":
        flash("Nahrajte prosím soubor s číselníkem katastrálních území.")
        return redirect(url_for("main.index"))

    try:
        kraje = parse_statistika(f_statistika.stream)
        population = parse_obyvatele(f_obyvatele)
        ratings = compute_ratings(kraje, population)
    except Exception as exc:
        flash(f"Chyba při zpracování souboru se statistikou nebo počtu obyvatel: {exc}")
        return redirect(url_for("main.index"))

    filename2 = f_ciselnik.filename or ""
    is_xlsx = filename2.lower().endswith((".xlsx", ".xlsm"))

    try:
        if is_xlsx:
            out_stream, out_name, nenaparovane, ku_kod_to_rating = process_xlsx_ciselnik(f_ciselnik.stream, ratings)
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            out_stream, out_name, nenaparovane, ku_kod_to_rating = process_csv_ciselnik(f_ciselnik.read(), ratings)
            mimetype = "text/csv"
    except Exception as exc:
        flash(f"Chyba při zpracování číselníku: {exc}")
        return redirect(url_for("main.index"))

    if nenaparovane:
        flash(
            "Pozor, tyto kraje z číselníku se nepodařilo napárovat na statistiku "
            "(sloupec Rating likvidity zůstal prázdný): " + ", ".join(sorted(nenaparovane))
        )

    if not ma_aktualizaci:
        return send_file(out_stream, as_attachment=True, download_name=out_name, mimetype=mimetype)

    try:
        akt_stream, akt_name, akt_nenaparovane, akt_pocet = process_aktualizace(f_aktualizace, ku_kod_to_rating)
    except Exception as exc:
        flash(f"Chyba při zpracování souboru k aktualizaci: {exc}")
        return redirect(url_for("main.index"))

    if akt_nenaparovane:
        flash(
            f"V souboru k aktualizaci se u {len(akt_nenaparovane)} katastrálních území "
            "nepodařilo najít odpovídající KU_KOD v číselníku, jejich Rating likvidity "
            "zůstal beze změny."
        )
    flash(f"V souboru k aktualizaci bylo aktualizováno {akt_pocet} řádků.")

    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(out_name, out_stream.getvalue())
        zf.writestr(akt_name, akt_stream.getvalue())
    zip_stream.seek(0)

    return send_file(zip_stream, as_attachment=True, download_name="vysledky.zip", mimetype="application/zip")
